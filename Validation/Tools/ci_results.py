#Script to update CI results in an excel file
import os
#import csv
from openpyxl import Workbook, load_workbook


class CI_Results(object):
    def __init__(self):
        self.ci_results_excel_file = "ci_results.xlsx"
        self.ci_file_obj = None
        self.ci_file = None
        self.pr_link = self.__get_pr_link()
        self.workflow_link = self.__get_workflow_link()
        self.pylint_result = self.__get_pylint_result()
        self.system_test_result = self.__get_system_test_result()
        self.overall_result = 'PASS' #default result
        self.failure_reason = 'NA' #default value in case everything passed


    def __get_pr_link(self):
        return os.getenv('PR_LINK')


    def __get_workflow_link(self):
        gh_repo = os.getenv('GITHUB_REPO')
        gh_repo_owner = os.getenv('GITHUB_REPO_OWNER')
        gh_run_id = os.getenv('GITHUB_RUN_ID')
        link = "https://github.com/{}/{}/actions/runs/{}".format(gh_repo, gh_repo_owner, gh_run_id)
        return link


    def __get_pylint_result(self):
        return os.getenv('PYLINT_RESULT')


    def __get_system_test_result(self):
        return os.getenv('SYSTEM_TEST_RESULT')


    def display_ci_results(self):
        print("Displaying Results")
        print("PR: {}, WORKFLOW_LINK: {}, PYLINT: {}, SYSTEM_TEST: {}".format(self.pr_link, self.workflow_link, self.pylint_result, self.system_test_result))


    def __get_excel_file(self):
        if os.path.exists(self.ci_results_excel_file):
            self.ci_file_obj = load_workbook(self.ci_results_excel_file)
            self.ci_file = self.ci_file_obj.active
        else:
            #create new file and add row attributes
            self.ci_file_obj = Workbook()
            self.ci_file = self.ci_file_obj.active
            self.ci_file.append(["PR_LINK", "WORKFLOW_LINK", "RESULT", "FAILURE_REASON"])

    def updateExcel(self):
        self.display_ci_results() #only for debug purpose
        self.__get_excel_file()

        #Update overall result & failure reason if any of the CI workflow step failed.
        if self.pylint_result == 'failed':
            self.overall_result = 'FAIL'
            self.failure_reason = 'pylint'
        elif self.system_test_result == 'failed':
            self.overall_result = 'FAIL'
            self.failure_reason = 'system_test'
        else:
            print("All CI checks passed")
        
        #Update excel
        self.ci_file.append([self.pr_link, self.workflow_link, self.overall_result, self.failure_reason])

        #save excel
        self.ci_file_obj.save(self.ci_results_excel_file)

    def displayExcelContent(self):
        print("Display CI results from excel")
        for row in self.ci_file.iter_rows(values_only=True):
            print(row)


if __name__ == '__main__':
    ci_obj = CI_Results()
    ci_obj.updateExcel()

    print("Results updated")
